#!/usr/bin/env python3
"""
Recommendation System for Online Marketplace
Full implementation with terminal interface - FIXED VERSION
"""

import json
import hashlib
import os
import random
import threading
import time
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
import openpyxl


AGE_MODIFIERS = {
    "0-13": {
        "Children's Products": 1.0,
        "Hobbies and Creativity": 0.6,
        "Gaming": 0.5,
        "Electronics": 0.2, 
        "Clothing and Footwear": 0.3,
        "Toys": 0.1,
        "Sports and Health": 0.15,
        "Perfume": 0.05,
        "Food and Goods": 0.1,
        "Auto Products": 0.05,
        "Pet Products": 0.2,
        "Lighting": 0.05,
        "Kitchen Products": 0.05,
        "Accessories": 0.1,
        "Home and Living": 0.05,
        "Tools and Repair": 0.05
    },
    "14-18": {
        "Children's Products": 0.1,
        "Hobbies and Creativity": 0.8,
        "Gaming": 0.9,
        "Electronics": 0.9,
        "Clothing and Footwear": 1.0,
        "Toys": 0.05,
        "Sports and Health": 0.7,
        "Perfume": 0.4,
        "Food and Goods": 0.5,
        "Auto Products": 0.2,
        "Pet Products": 0.3,
        "Lighting": 0.2,
        "Kitchen Products": 0.1,
        "Accessories": 0.9,
        "Home and Living": 0.1,
        "Tools and Repair": 0.1
    },
    "19-30": {
        "Children's Products": 0.4,
        "Hobbies and Creativity": 0.6,
        "Gaming": 0.7,
        "Electronics": 1.0,
        "Clothing and Footwear": 1.0,
        "Toys": 0.1,
        "Sports and Health": 0.9,
        "Perfume": 0.8,
        "Food and Goods": 0.6,
        "Auto Products": 0.7,
        "Pet Products": 0.4,
        "Lighting": 0.5,
        "Kitchen Products": 0.6,
        "Accessories": 1.0,
        "Home and Living": 0.7,
        "Tools and Repair": 0.5
    },
    "31-50": {
        "Children's Products": 0.4,
        "Hobbies and Creativity": 0.5,
        "Gaming": 0.2,
        "Electronics": 0.9,
        "Clothing and Footwear": 0.6,
        "Toys": 0.2,
        "Sports and Health": 0.9,
        "Perfume": 0.7,
        "Food and Goods": 1.0,
        "Auto Products": 1.0,
        "Pet Products": 0.8,
        "Lighting": 0.8,
        "Kitchen Products": 0.8,
        "Accessories": 0.7,
        "Home and Living": 1.0,
        "Tools and Repair": 0.9
    },
    "50+": {
        "Children's Products": 0.2,
        "Hobbies and Creativity": 0.8,
        "Gaming": 0.2,
        "Electronics": 0.6,
        "Clothing and Footwear": 0.7,
        "Toys": 0.2,
        "Sports and Health": 0.7,
        "Perfume": 0.6,
        "Food and Goods": 0.9,
        "Auto Products": 0.8,
        "Pet Products": 0.7,
        "Lighting": 0.9,
        "Kitchen Products": 1.0,
        "Accessories": 0.6,
        "Home and Living": 1.0,
        "Tools and Repair": 1.0
    }
}

GENDER_MODIFIERS = {
    "male": {
        "0-13": {"enhance": [], "reduce": []},
        "14-18": {"enhance": [], "reduce": []},
        "19-30": {"enhance": [], "reduce": []},
        "31-50": {"enhance": [], "reduce": []},
        "50+": {"enhance": [], "reduce": []}
    },
    "female": {
        "0-13": {"enhance": [], "reduce": []},
        "14-18": {"enhance": [], "reduce": []},
        "19-30": {"enhance": [], "reduce": []},
        "31-50": {"enhance": [], "reduce": []},
        "50+": {"enhance": [], "reduce": []}
    }
}

LOCATION_MODIFIERS = {
    "big_city": {
        "premium": 1.0,
        "medium": 0.6,
        "budget": 0.3,  # Increased from 0.1
        "expensive": 1.0,
        "average": 0.6,
        "cheap": 0.3,  # Increased from 0.1
        "1day": 1.0,
        "2-3days": 0.5,
        "4+days": 0.3  # Increased from 0.1
    },
    "small_city": {
        "premium": 0.6,
        "medium": 1.0,
        "budget": 0.5,  # Increased from 0.3
        "expensive": 0.6,
        "average": 1.0,
        "cheap": 0.5,  # Increased from 0.3
        "1day": 0.5,
        "2-3days": 1.0,
        "4+days": 0.5  # Increased from 0.3
    },
    "village": {
        "premium": 0.3,  # Increased from 0.2
        "medium": 0.6,
        "budget": 1.0,
        "expensive": 0.3,  # Increased from 0.2
        "average": 0.6,
        "cheap": 1.0,
        "1day": 0.3,  # Increased from 0.2
        "2-3days": 0.6,
        "4+days": 1.0
    }
}

LOCATION_SPHERE_MODIFIERS = {
    "big_city": {
        "enhance": ["Electronics", "Clothing and Footwear", "Perfume", "Gaming", "Hobbies and Creativity"],
        "reduce": ["Auto Products", "Tools and Repair"]
    },
    "small_city": {
        "enhance": ["Electronics", "Clothing and Footwear", "Home and Living"],
        "reduce": []
    },
    "village": {
        "enhance": ["Tools and Repair", "Auto Products", "Home and Living", "Kitchen Products", "Food and Goods"],
        "reduce": ["Perfume", "Gaming", "Accessories"]
    }
}

TAG_CATEGORIES = {
    "lifestyle": ["premium", "budget", "eco-friendly", "minimalist", "luxury"],
    "tech": ["wireless", "wired", "smart", "manual", "automatic", "portable"],
    "activity": ["outdoor", "indoor", "travel", "home", "office", "gym"],
    "aesthetic": ["modern", "classic", "vintage", "colorful", "monochrome"],
    "gaming": ["rpg", "fps", "strategy", "casual", "multiplayer"],
    "sports": ["running", "gym", "yoga", "cycling", "swimming"],
    "clothing": ["casual", "formal", "sportswear", "streetwear"],
}

SPHERE_TYPES = {
    "Tools and Repair": ["Drills", "Screwdrivers", "Hammers", "Saws", "Wrenches", "Pliers", "Sanders", "Grinders", "Measuring Tools", "Power Tools", "Hand Tools", "Cutting Tools", "Assembly Tools"],
    "Gaming": ["Gaming Consoles", "Gaming Mice", "Gaming Keyboards", "Gaming Headsets", "Gaming Monitors", "Gaming Chairs", "Controllers", "Gamepad", "VR Equipment", "Gaming Desk", "Fight Sticks"],
    "Sports and Health": ["Dumbbells", "Treadmills", "Yoga Mats", "Protein Powder", "Running Shoes", "Bicycles", "Swimming Gear", "Gym Equipment", "Fitness Tracker", "Resistance Bands", "Jump Rope", "Boxing Gloves", "Kettlebells"],
    "Clothing and Footwear": ["Jackets", "Sneakers", "Jeans", "T-Shirts", "Dresses", "Hoodies", "Shorts", "Boots", "Sandals", "Socks", "Underwear", "Sportswear", "Formal Wear", "Winter Coat", "Athletic Shoes"],
    "Perfume": ["Men's Cologne", "Women's Perfume", "Unisex Fragrance", "Body Spray", "Aftershave", "Deodorant", "Fragrance Gift Set"],
    "Food and Goods": ["Coffee", "Tea", "Chocolate", "Snacks", "Spices", "Condiments", "Cereal", "Pasta", "Canned Goods", "Beverages", "Nuts", "Dried Fruits"],
    "Auto Products": ["Tires", "Oil", "Air Filter", "Brake Pads", "Car Batteries", "Windshield Wipers", "Spark Plugs", "Car Seats", "Floor Mats", "Roof Rack"],
    "Pet Products": ["Dog Food", "Cat Food", "Pet Toys", "Pet Bed", "Leash", "Collar", "Pet Cage", "Fish Tank", "Pet Treats", "Grooming Supplies", "Pet Carrier"],
    "Hobbies and Creativity": ["Paints", "Brushes", "Canvas", "Sketchbook", "Colored Pencils", "Markers", "Clay", "Knitting Needles", "Yarn", "Photography Equipment", "Musical Instrument", "Craft Kit"],
    "Lighting": ["Lamps", "Chandeliers", "LED Bulbs", "Desk Lamp", "Floor Lamp", "Wall Sconce", "String Lights", "Flashlight", "Lantern", "Smart Lights"],
    "Kitchen Products": ["Pans", "Knives", "Blenders", "Coffee Maker", "Microwave", "Toaster", "Cutting Board", "Measuring Cups", "Pot Set", "Utensils", "Dish Set", "Baking Tools"],
    "Children's Products": ["Toys", "Building Blocks", "Educational Games", "Stroller", "Car Seat", "Baby Monitor", "Crib", "Playpen", "Diapers", "Baby Clothes", "Action Figures"],
    "Accessories": ["Watches", "Bags", "Belts", "Scarves", "Hats", "Gloves", "Sunglasses", "Jewelry", "Wallets", "Phone Cases", "Backpack", "Keychain"],
    "Electronics": ["Smartphones", "Laptops", "Tablets", "Televisions", "Cameras", "Headphones", "Smart Watch", "Charger", "Power Bank", "USB Cable", "Router", "Monitor", "Keyboard"],
    "Home and Living": ["Sofas", "Beds", "Tables", "Chairs", "Shelves", "Cabinets", "Nightstand", "Bookcase", "Mirrors", "Rugs", "Curtains", "Bedding", "Pillows"],
}

SPHERE_MAPPING = {
    "Инструменты и ремонт": "Tools and Repair",
    "Гейминг": "Gaming",
    "Спорт и здоровье": "Sports and Health",
    "Одежда и обувь": "Clothing and Footwear",
    "Парфюм": "Perfume",
    "Еда и товары": "Food and Goods",
    "Автротовары": "Auto Products",
    "Товары для животных": "Pet Products",
    "Хобби и творчество": "Hobbies and Creativity",
    "Освещение": "Lighting",
    "Кухонные товары": "Kitchen Products",
    "Детские товары": "Children's Products",
    "Аксессуары": "Accessories",
    "Техника и элэктроника": "Electronics",
    "Дом и быт": "Home and Living"
}

QUALITY_CRITERIA = ["premium", "medium", "budget"]
PRICE_CRITERIA = ["expensive", "average", "cheap"]
DELIVERY_CRITERIA = ["1day", "2-3days", "4+days"]

SPHERE_DECAY_CONFIG = {
    "next_decay_time": {},
    "cancel_decay_flags": {},
    "base_interval": 20 * 60,
    "purchase_interval": 10 * 60,
    "decay_lock": threading.Lock(),
    "background_thread": None,
    "products_ref": None,
}

MAX_SPHERE_SCORE = 5.0
MAX_CRITERIA_SCORE = 3.0
MAX_TAG_SCORE = 3.0
MAX_TYPE_SCORE = 3.0


def get_seasonal_bonus(sphere: str) -> float:
    """
    FIXED: Increased seasonal effects for more noticeable impact
    """
    month = datetime.now().month
    
    if month in [12, 1, 2]:
        if sphere == "Clothing and Footwear":
            return 1.3  # Increased from 1.2
        if sphere == "Sports and Health":
            return 0.7  # Decreased from 0.9 - stronger effect
    
    elif month in [6, 7, 8]:
        if sphere == "Sports and Health":
            return 1.5  # Increased from 1.3 - stronger boost
        if sphere == "Clothing and Footwear":
            return 1.2  # Increased from 1.1
        if sphere == "Gaming":
            return 0.75  # Decreased from 0.85 - people outside more
    
    return 1.0


def check_sphere_decay(products: List[Dict], sphere: str) -> bool:
    """
    FIXED: Corrected decay formula logic
    """
    with SPHERE_DECAY_CONFIG["decay_lock"]:
        current_time = datetime.now().timestamp()
        next_decay = SPHERE_DECAY_CONFIG["next_decay_time"].get(sphere, 0)
        
        if current_time < next_decay:
            return False
        
        if SPHERE_DECAY_CONFIG["cancel_decay_flags"].get(sphere, False):
            SPHERE_DECAY_CONFIG["cancel_decay_flags"][sphere] = False
            SPHERE_DECAY_CONFIG["next_decay_time"][sphere] = current_time + SPHERE_DECAY_CONFIG["purchase_interval"]
            return False
        
        sphere_products = [p for p in products if p.get("sphere") == sphere]
        
        if len(sphere_products) < 3:
            SPHERE_DECAY_CONFIG["next_decay_time"][sphere] = current_time + SPHERE_DECAY_CONFIG["base_interval"]
            return False
        
        sphere_products_sorted = sorted(sphere_products, key=lambda p: p.get("_score", 0), reverse=True)
        product_3rd = sphere_products_sorted[2]
        
        if "decay_score" not in product_3rd:
            product_3rd["decay_score"] = 1.0
        
        if SPHERE_DECAY_CONFIG["cancel_decay_flags"].get(sphere, False):
            SPHERE_DECAY_CONFIG["cancel_decay_flags"][sphere] = False
            SPHERE_DECAY_CONFIG["next_decay_time"][sphere] = current_time + SPHERE_DECAY_CONFIG["purchase_interval"]
            return False
        
        old_score = product_3rd["decay_score"]
        new_score = max(old_score * 0.9, 0.3)  # Never go below 0.3
        product_3rd["decay_score"] = new_score
        
        SPHERE_DECAY_CONFIG["next_decay_time"][sphere] = current_time + SPHERE_DECAY_CONFIG["base_interval"]
        return True


def cancel_sphere_decay(sphere: str) -> None:
    """Cancel next decay for sphere on purchase"""
    with SPHERE_DECAY_CONFIG["decay_lock"]:
        current_time = datetime.now().timestamp()
        SPHERE_DECAY_CONFIG["cancel_decay_flags"][sphere] = True
        SPHERE_DECAY_CONFIG["next_decay_time"][sphere] = current_time + SPHERE_DECAY_CONFIG["purchase_interval"]


def start_decay_background_task(products: List[Dict], all_spheres: List[str]) -> None:
    """Start background decay process"""
    SPHERE_DECAY_CONFIG["products_ref"] = products
    
    def background_decay_loop():
        while True:
            time.sleep(10)
            
            for sphere in all_spheres:
                if SPHERE_DECAY_CONFIG["next_decay_time"].get(sphere, 0) <= datetime.now().timestamp():
                    check_sphere_decay(products, sphere)
    
    thread = threading.Thread(target=background_decay_loop, daemon=True)
    thread.start()
    SPHERE_DECAY_CONFIG["background_thread"] = thread


def normalize_criteria_group(user: 'User', criteria_list: List[str]):
    """
    IMPROVED: Softer normalization with higher minimum threshold
    """
    total = sum(user.criteria_scores.get(c, 0.1) for c in criteria_list)
    
    if total > 3.0:  # Increased threshold from 2.0
        factor = 3.0 / total  # Normalize to 3.0 instead of 2.0
        for criterion in criteria_list:
            if criterion in user.criteria_scores:
                user.criteria_scores[criterion] *= factor
                user.criteria_scores[criterion] = max(0.2, min(user.criteria_scores[criterion], MAX_CRITERIA_SCORE))
    else:
        for criterion in criteria_list:
            if criterion in user.criteria_scores:
                user.criteria_scores[criterion] = max(0.2, user.criteria_scores[criterion])


class Database:
    """Simple JSON-based database for users and products"""
    
    def __init__(self):
        self.users_file = "users.json"
        self.products_file = "products.json"
        self.transactions_file = "transactions.json"
        self._init_files()
    
    def _init_files(self):
        """Initialize database files if they don't exist"""
        for file in [self.users_file, self.products_file, self.transactions_file]:
            if not os.path.exists(file):
                with open(file, 'w') as f:
                    json.dump({}, f)
    
    def load_users(self):
        """FIXED: Added type checking for loaded data"""
        try:
            with open(self.users_file, "r") as f:
                data = f.read().strip()
                if not data:
                    return {}  
                result = json.loads(data)
                if not isinstance(result, dict):
                    return {}
                return result
        except FileNotFoundError:
            return {}          
        except json.JSONDecodeError:
            return {}
    
    def save_users(self, users: Dict):
        """Save all users to database"""
        with open(self.users_file, "w") as f:
            json.dump(users, f, indent=2, ensure_ascii=False)
    
    def load_products(self) -> List[Dict]:
        """Load all products from database"""
        with open(self.products_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, dict):
                return data.get("products", [])
            elif isinstance(data, list):
                return data
            else:
                return []
    
    def save_products(self, products: List[Dict]):
        """Save all products to database"""
        with open(self.products_file, 'w', encoding='utf-8') as f:
            json.dump({"products": products}, f, indent=2, ensure_ascii=False)
    
    def load_transactions(self) -> List[Dict]:
        """Load all transactions"""
        with open(self.transactions_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get("transactions", [])
    
    def save_transaction(self, transaction: Dict):
        """Save a new transaction"""
        transactions = self.load_transactions()
        transactions.append(transaction)
        with open(self.transactions_file, 'w', encoding='utf-8') as f:
            json.dump({"transactions": transactions}, f, indent=2, ensure_ascii=False)


UNIVERSAL_TAGS = {
    "lifestyle": ["premium", "budget", "eco-friendly", "luxury", "affordable", "value"],
    "usage": ["daily", "occasional", "professional", "casual", "formal"],
    "quality": ["durable", "long-lasting", "sturdy", "reliable"],
    "convenience": ["portable", "compact", "lightweight", "handheld", "foldable", "easy-to-use"],
    "tech": ["wireless", "smart", "digital", "usb", "rechargeable", "battery-powered"],
    "activity": ["indoor", "outdoor", "travel", "home", "office"],
    "gift": ["gift-ready", "collectible", "showpiece"],
    "aesthetic": ["modern", "classic", "stylish", "elegant"],
}

PRODUCT_TAGS = {
    "Дрели": ["tools", "construction", "diy", "power-tool", "professional", "durable"],
    "Болгарки": ["tools", "construction", "cutting", "grinding", "professional", "powerful"],
    "Отвертки": ["tools", "hand-tool", "assembly", "portable", "compact", "affordable"],
    
    "Игровые консоли": ["gaming", "entertainment", "home", "multiplayer", "premium"],
    "Игровые мышки": ["gaming", "tech", "wireless", "portable", "precision", "professional"],
    "Игровые клавиатуры": ["gaming", "tech", "wireless", "portable", "professional", "stylish"],
    
    "Гантели": ["fitness", "strength", "home", "training", "compact", "durable"],
    "Тренажёры": ["fitness", "cardio", "home", "training", "powerful", "sturdy"],
    "Йога маты": ["fitness", "yoga", "home", "portable", "lightweight", "foldable"],
    
    "Куртки": ["clothing", "outerwear", "weather-protection", "casual", "stylish", "durable"],
    "Кроссовки": ["footwear", "sports", "casual", "comfortable", "active", "travel"],
    "Джинсы": ["clothing", "casual", "everyday", "durable", "stylish", "comfortable"],
    
    "Мужские духи": ["fragrance", "personal-care", "luxury", "gift-ready", "stylish"],
    "Женские духи": ["fragrance", "personal-care", "luxury", "gift-ready", "premium"],
    
    "Кофе": ["beverages", "daily", "home", "affordable", "quality"],
    "Чай": ["beverages", "daily", "home", "relaxation", "healthy"],
    "Шоколад": ["snacks", "gift-ready", "indulgence", "affordable", "collectible"],
    
    "Шины": ["automotive", "safety", "durable", "essential", "powerful", "professional"],
    "Масло": ["automotive", "maintenance", "essential", "affordable", "quality"],
    
    "Корм для собак": ["pets", "daily", "essential", "healthy", "quality", "home"],
    "Корм для кошек": ["pets", "daily", "essential", "healthy", "quality", "home"],
    "Игрушки": ["pets", "entertainment", "interactive", "home", "affordable"],
    
    "Краски": ["art", "creative", "hobby", "professional", "quality", "stylish"],
    "Кисти": ["art", "creative", "hobby", "professional", "precision", "durable"],
    "Холсты": ["art", "creative", "hobby", "professional", "quality", "stylish"],
    
    "Лампы": ["lighting", "home", "modern", "energy-efficient", "essential"],
    "Люстры": ["lighting", "home-decor", "elegant", "stylish", "premium", "showpiece"],
    
    "Сковородки": ["kitchen", "cooking", "essential", "daily", "durable", "quality"],
    "Ножи": ["kitchen", "cooking", "essential", "sharp", "durable", "professional"],
    "Блендеры": ["kitchen", "appliance", "convenience", "healthy", "daily", "portable"],
    
    "Игрушки детские": ["kids", "toys", "entertainment", "creative", "educational"],
    "Конструкторы": ["kids", "toys", "educational", "creative", "gift-ready"],
    
    "Часы": ["accessories", "stylish", "luxury", "gift-ready", "timekeeping"],
    "Сумки": ["accessories", "stylish", "practical", "travel", "everyday", "durable"],
    "Ремни": ["accessories", "stylish", "practical", "everyday", "durable"],
    
    "Смартфоны": ["mobile", "tech", "wireless", "portable", "premium", "daily"],
    "Ноутбуки": ["computer", "tech", "portable", "work", "productivity", "premium"],
    "Телевизоры": ["entertainment", "home", "tech", "large-screen", "stylish", "premium"],
    
    "Диваны": ["furniture", "home", "comfort", "luxury", "stylish", "durable"],
    "Кровати": ["furniture", "home", "comfort", "essential", "durable", "quality"],
    "Столы": ["furniture", "home", "work", "functional", "modern", "durable"],
}


def load_products_from_excel(filepath: str) -> List[Dict]:
    """Load products from Excel file and convert to product objects"""
    wb = openpyxl.load_workbook(filepath)
    products = []
    product_id = 1
    
    for sheet_name in wb.sheetnames:
        sphere = SPHERE_MAPPING.get(sheet_name, sheet_name)
        sheet = wb[sheet_name]
        
        current_row = 1
        while current_row <= sheet.max_row:
            cell_value = sheet.cell(current_row, 1).value
            if cell_value and isinstance(cell_value, str) and cell_value.strip():
                product_type = cell_value.strip()
                
                if current_row + 1 <= sheet.max_row:
                    criteria_row = current_row + 1
                    
                    data_row = current_row + 2
                    while data_row <= sheet.max_row:
                        next_cell = sheet.cell(data_row, 1).value
                        if next_cell and isinstance(next_cell, str) and not any(
                            sheet.cell(data_row, col).value for col in range(2, 10)
                        ):
                            break
                        
                        product_name = sheet.cell(data_row, 1).value
                        if product_name and isinstance(product_name, str):
                            product = {
                                "id": product_id,
                                "name": product_name.strip(),
                                "sphere": sphere,
                                "type": product_type,
                                "price": round(random.uniform(10, 500), 2),
                                "owner": "system",
                                "quality": None,
                                "price_level": None,
                                "delivery": None,
                                "tags": []
                            }
                            
                            for col in range(1, 4):
                                if sheet.cell(data_row, col).value:
                                    if col == 1:
                                        product["quality"] = "premium"
                                    elif col == 2:
                                        product["quality"] = "medium"
                                    elif col == 3:
                                        product["quality"] = "budget"
                                    break
                            
                            for col in range(4, 7):
                                if sheet.cell(data_row, col).value:
                                    if col == 4:
                                        product["price_level"] = "cheap"
                                    elif col == 5:
                                        product["price_level"] = "average"
                                    elif col == 6:
                                        product["price_level"] = "expensive"
                                    break
                            
                            for col in range(7, 10):
                                if sheet.cell(data_row, col).value:
                                    if col == 7:
                                        product["delivery"] = "1day"
                                    elif col == 8:
                                        product["delivery"] = "2-3days"
                                    elif col == 9:
                                        product["delivery"] = "4+days"
                                    break
                            
                            tags_cell = sheet.cell(data_row, 10).value
                            if tags_cell and isinstance(tags_cell, str):
                                product["tags"] = [t.strip() for t in tags_cell.split(',') if t.strip()]
                            else:
                                if product_type in PRODUCT_TAGS:
                                    product["tags"] = PRODUCT_TAGS[product_type]
                                else:
                                    product["tags"] = ["general", "product"]
                            
                            if all([product["quality"], product["price_level"], product["delivery"]]):
                                products.append(product)
                                product_id += 1
                        
                        data_row += 1
                    
                    current_row = data_row
                else:
                    current_row += 1
            else:
                current_row += 1
    
    return products


class User:
    """User class with profile and preferences"""
    
    def __init__(self, username: str, password: str, age: int, gender: str, location: str, balance: int):
        self.username = username
        self.password_hash = hashlib.sha256(password.encode()).hexdigest()
        self.age = age
        self.gender = gender
        self.location = location
        self.balance = max(0, balance)
        
        age_group = self._get_age_group()
        self.sphere_scores = self._init_sphere_scores(age_group, gender)
        self.criteria_scores = self._init_criteria_scores(location)
        
        self.tag_scores = {}
        self.type_scores = {}
        
        self.initial_influence = 1.0
        
        self.last_purchase_date = {}
        self.recommended_purchases = set()
        self.purchase_history = []
    
    def _get_age_group(self) -> str:
        """Determine age group"""
        if self.age <= 13:
            return "0-13"
        elif self.age <= 18:
            return "14-18"
        elif self.age <= 30:
            return "19-30"
        elif self.age <= 50:
            return "31-50"
        else:
            return "50+"
    
    def _init_sphere_scores(self, age_group: str, gender: str) -> Dict[str, float]:
        """Initialize sphere scores based on age, gender, and location"""
        scores = AGE_MODIFIERS[age_group].copy()
        
        gender_mods = GENDER_MODIFIERS[gender][age_group]
        for sphere in gender_mods["enhance"]:
            if sphere in scores:
                scores[sphere] *= 1.2
        
        for sphere in gender_mods["reduce"]:
            if sphere in scores and sphere not in gender_mods["enhance"]:
                scores[sphere] *= 0.85
        
        location_mods = LOCATION_SPHERE_MODIFIERS.get(self.location, {})
        for sphere in location_mods.get("enhance", []):
            if sphere in scores:
                scores[sphere] *= 1.1
        
        for sphere in location_mods.get("reduce", []):
            if sphere in scores and sphere not in location_mods.get("enhance", []):
                scores[sphere] *= 0.9
        
        for sphere in scores:
            scores[sphere] = max(scores[sphere], 0.1)
        
        return scores
    
    def _init_criteria_scores(self, location: str) -> Dict[str, float]:
        """Initialize criteria scores based on location"""
        return LOCATION_MODIFIERS[location].copy()
    
    def to_dict(self) -> Dict:
        """Convert user to dictionary for storage"""
        return {
            "username": self.username,
            "password_hash": self.password_hash,
            "age": self.age,
            "gender": self.gender,
            "location": self.location,
            "balance": self.balance,
            "sphere_scores": self.sphere_scores,
            "criteria_scores": self.criteria_scores,
            "tag_scores": self.tag_scores,
            "type_scores": self.type_scores,
            "initial_influence": self.initial_influence,
            "last_purchase_date": self.last_purchase_date,
            "recommended_purchases": list(self.recommended_purchases),
            "purchase_history": self.purchase_history
        }
    
    @staticmethod
    def from_dict(data: Dict) -> 'User':
        """Create user from dictionary"""
        user = object.__new__(User)
        user.username = data["username"]
        user.password_hash = data["password_hash"]
        user.age = data["age"]
        user.gender = data["gender"]
        user.location = data["location"]
        user.balance = data["balance"]
        user.sphere_scores = data["sphere_scores"]
        user.criteria_scores = data["criteria_scores"]
        user.tag_scores = data.get("tag_scores", {})
        user.type_scores = data.get("type_scores", {})
        user.initial_influence = data["initial_influence"]
        user.last_purchase_date = data.get("last_purchase_date", {})
        user.recommended_purchases = set(data.get("recommended_purchases", []))
        user.purchase_history = data.get("purchase_history", [])
        return user


class RecommendationEngine:
    """Core recommendation algorithm"""
    
    @staticmethod
    def calculate_product_score(user: User, product: Dict) -> float:
        """Calculate final score for a product"""
        sphere_score = user.sphere_scores.get(product["sphere"], 0.1)
        
        sphere_score = sphere_score * (0.3 + user.initial_influence * 0.7)
        
        seasonal_bonus = get_seasonal_bonus(product["sphere"])
        sphere_score *= seasonal_bonus
        
        quality_score = user.criteria_scores.get(product["quality"], 0.1)
        price_score = user.criteria_scores.get(product["price_level"], 0.1)
        delivery_score = user.criteria_scores.get(product["delivery"], 0.1)
        
        tag_score = 0.1
        if product.get("tags"):
            tag_values = [user.tag_scores.get(tag, 0.1) for tag in product["tags"]]
            tag_values.sort(reverse=True)
            if len(tag_values) >= 2:
                tag_score = (tag_values[0] + tag_values[1]) / 2
            elif len(tag_values) == 1:
                tag_score = tag_values[0]
        
        type_score = user.type_scores.get(product["type"], 0.1)
        
        final_score = (
            sphere_score * 0.35 +
            quality_score * 0.15 +
            price_score * 0.15 +
            delivery_score * 0.10 +
            tag_score * 0.15 +
            type_score * 0.10
        )
        
        return final_score
    
    @staticmethod
    def get_recommendations(user: User, products: List[Dict], count: int = 30) -> List[Dict]:
        """Get personalized recommendations - interleaved by sphere for perfect balance"""
        scored_products = []
        
        for p in products:
            base_score = RecommendationEngine.calculate_product_score(user, p)
            
            sphere = p["sphere"]
            last_purchase = user.last_purchase_date.get(sphere)
            
            if last_purchase:
                days_ago = (datetime.now() - datetime.fromisoformat(last_purchase)).days
                if days_ago > 30:
                    base_score *= 1.10
            else:
                base_score *= 1.15
            
            if "decay_score" in p:
                base_score *= p["decay_score"]
            
            p["_score"] = base_score
            scored_products.append((p, base_score))
        
        scored_products.sort(key=lambda x: x[1], reverse=True)
        
        sorted_spheres = sorted(
            user.sphere_scores.items(),
            key=lambda x: x[1],
            reverse=True
        )
        top_spheres = [s[0] for s in sorted_spheres[:5]]
        
        products_by_sphere = {}
        for sphere in top_spheres:
            products_by_sphere[sphere] = [p for p, s in scored_products if p["sphere"] == sphere]
        
        other_products = [p for p, s in scored_products if p["sphere"] not in top_spheres]
        
        recommendations = []
        sphere_indices = {s: 0 for s in top_spheres}
        max_per_sphere = (count // len(top_spheres)) + 2
        
        sphere_queue = top_spheres.copy()
        other_idx = 0
        
        while len(recommendations) < count:
            added_in_round = False
            
            active_spheres = [s for s in sphere_queue 
                            if sphere_indices[s] < len(products_by_sphere[s]) 
                            and sphere_indices[s] < max_per_sphere]
            
            if not active_spheres and other_idx >= len(other_products):
                break
            
            for sphere in active_spheres:
                if len(recommendations) >= count:
                    break
                
                product = products_by_sphere[sphere][sphere_indices[sphere]]
                recommendations.append(product)
                sphere_indices[sphere] += 1
                added_in_round = True
            
            if not added_in_round and other_idx < len(other_products):
                recommendations.append(other_products[other_idx])
                other_idx += 1
        
        return recommendations[:count]
    
    @staticmethod
    def update_profile_after_purchase(user: User, product: Dict, was_recommended: bool):
        """
        FIXED: Added hard caps to prevent infinite score growth
        """
        sphere = product["sphere"]
        
        if was_recommended:
            increase = 0.15
        else:
            increase = 0.1
        
        user.sphere_scores[sphere] = min(
            user.sphere_scores[sphere] + increase,
            MAX_SPHERE_SCORE  # Hard cap at 5.0
        )
        
        quality = product["quality"]
        price_level = product["price_level"]
        delivery = product["delivery"]
        
        user.criteria_scores[quality] = min(
            user.criteria_scores[quality] + 0.1,
            MAX_CRITERIA_SCORE  # Hard cap at 3.0
        )
        user.criteria_scores[price_level] = min(
            user.criteria_scores[price_level] + 0.1,
            MAX_CRITERIA_SCORE
        )
        user.criteria_scores[delivery] = min(
            user.criteria_scores[delivery] + 0.1,
            MAX_CRITERIA_SCORE
        )
        
        for criterion in QUALITY_CRITERIA:
            if criterion != quality:
                user.criteria_scores[criterion] = max(0.2, user.criteria_scores[criterion] - 0.05)
        
        for criterion in PRICE_CRITERIA:
            if criterion != price_level:
                user.criteria_scores[criterion] = max(0.2, user.criteria_scores[criterion] - 0.05)
        
        for criterion in DELIVERY_CRITERIA:
            if criterion != delivery:
                user.criteria_scores[criterion] = max(0.2, user.criteria_scores[criterion] - 0.05)
        
        normalize_criteria_group(user, QUALITY_CRITERIA)
        normalize_criteria_group(user, PRICE_CRITERIA)
        normalize_criteria_group(user, DELIVERY_CRITERIA)
        
        if user.initial_influence > 0.2:
            age_group = user._get_age_group()
            
            if age_group == "0-13":
                decay = 0.98
            elif age_group == "14-18":
                decay = 0.92
            elif age_group == "19-30":
                decay = 0.90
            elif age_group == "31-50":
                decay = 0.93
            else:
                decay = 0.96
            
            user.initial_influence *= decay
        
        for s in user.sphere_scores:
            if s != sphere:
                user.sphere_scores[s] = max(0.3, user.sphere_scores[s] * 0.98)
        
        for tag in product.get("tags", []):
            if tag in user.tag_scores:
                user.tag_scores[tag] = min(user.tag_scores[tag] + 0.15, MAX_TAG_SCORE)
            else:
                user.tag_scores[tag] = 0.15
        
        product_type = product["type"]
        if product_type in user.type_scores:
            user.type_scores[product_type] = min(user.type_scores[product_type] + 0.3, MAX_TYPE_SCORE)
        else:
            user.type_scores[product_type] = 0.3
        
        user.last_purchase_date[sphere] = datetime.now().isoformat()
        
        cancel_sphere_decay(sphere)


class TerminalInterface:
    """Terminal-based user interface"""
    
    def __init__(self):
        self.db = Database()
        self.current_user = None
        self.products = []
        self.recommendations = []
    
    def clear_screen(self):
        """Clear terminal screen"""
        os.system('clear' if os.name != 'nt' else 'cls')
    
    def print_header(self, title: str):
        """Print section header"""
        print("\n" + "=" * 60)
        print(f"  {title}")
        print("=" * 60 + "\n")
    
    def print_menu(self, options: List[Tuple[int, str]]):
        """Print menu options"""
        for num, text in options:
            print(f"{num}. {text}")
        print()
    
    def get_choice(self, max_option: int) -> int:
        """Get user menu choice"""
        while True:
            try:
                choice = int(input("Enter your choice: "))
                if 1 <= choice <= max_option:
                    return choice
                print(f"Please enter a number between 1 and {max_option}")
            except ValueError:
                print("Please enter a valid number")
    
    def run(self):
        """Main application loop"""
        print("Loading products...")
        excel_path = "IA_COMP_EXPANDED.xlsx"
        if os.path.exists(excel_path):
            self.products = load_products_from_excel(excel_path)
            self.db.save_products(self.products)
            print(f"Loaded {len(self.products)} products")
        else:
            self.products = self.db.load_products()
            print(f"Loaded {len(self.products)} products from database")
        
        all_spheres = list(set(p.get("sphere", "") for p in self.products))
        start_decay_background_task(self.products, all_spheres)
        
        while True:
            self.clear_screen()
            self.print_header("MARKETPLACE - MAIN MENU")
            
            if self.current_user:
                print(f"Logged in as: {self.current_user.username}")
                print(f"Balance: ${self.current_user.balance:.2f}\n")
                self.print_menu([
                    (1, "View Recommendations"),
                    (2, "Browse All Products"),
                    (3, "View My Profile"),
                    (4, "View Purchase History"),
                    (5, "Add Product for Sale"),
                    (6, "Manage My Listings"),
                    (7, "Replenish Balance"),
                    (8, "Improve Recommendations"),
                    (9, "Logout"),
                    (10, "Exit")
                ])
                choice = self.get_choice(10)
                
                if choice == 1:
                    self.view_recommendations()
                elif choice == 2:
                    self.browse_products()
                elif choice == 3:
                    self.view_profile()
                elif choice == 4:
                    self.view_purchase_history()
                elif choice == 5:
                    self.add_product()
                elif choice == 6:
                    self.manage_listings()
                elif choice == 7:
                    self.replenish_balance()
                elif choice == 8:
                    self.improve_recommendations()
                elif choice == 9:
                    self.logout()
                elif choice == 10:
                    break
            else:
                self.print_menu([
                    (1, "Register"),
                    (2, "Login"),
                    (3, "Exit")
                ])
                choice = self.get_choice(3)
                
                if choice == 1:
                    self.register()
                elif choice == 2:
                    self.login()
                elif choice == 3:
                    break
        
        print("\nThank you for using the Marketplace! Goodbye!")
    
    def improve_recommendations(self):
        """FIXED: Reduced boost from 0.5 to 0.15"""
        self.clear_screen()
        self.print_header("IMPROVE RECOMMENDATIONS")

        print("Choose your age group:")
        print("1. Under 18")
        print("2. 18 or older")
        choice = self.get_choice(2)

        if choice == 1:
            self.questionnaire_under_18()
        else:
            self.questionnaire_over_18()

    def questionnaire_under_18(self):
        """FIXED: Reduced sphere boost"""
        self.clear_screen()
        self.print_header("QUESTIONNAIRE - UNDER 18")

        questions = [
            ("Do you need something for school or studying?", ["Children's Products", "Accessories"]),
            ("Are you looking for a toy or board game?", ["Toys", "Gaming"]),
            ("Do you want something for creativity?", ["Hobbies and Creativity"]),
            ("Are you choosing something for your computer or phone?", ["Electronics"]),
            ("Are you looking for clothing or shoes?", ["Clothing and Footwear"]),
            ("Do you need something for sports or activities?", ["Sports and Health"]),
            ("Are you choosing a gift for a friend?", ["Toys", "Accessories", "Gaming"]),
            ("Do you need something for your room—decor or lighting?", ["Home and Living", "Lighting"]),
            ("Do you want something for pets?", ["Pet Products"]),
            ("Do you need something for the kitchen or home?", ["Kitchen Products", "Home and Living"]),
            ("Do you want something for playing outside?", ["Toys", "Sports and Health"]),
            ("Are you searching for hobby items like craft sets or models?", ["Hobbies and Creativity"]),
        ]

        self.process_questionnaire(questions)

    def questionnaire_over_18(self):
        """FIXED: Reduced sphere boost"""
        self.clear_screen()
        self.print_header("QUESTIONNAIRE - 18 OR OLDER")

        questions = [
            ("Are you looking for something for your home?", ["Home and Living", "Lighting", "Kitchen Products"]),
            ("Do you need a tool or something for repair?", ["Tools and Repair"]),
            ("Are you choosing electronics or tech?", ["Electronics"]),
            ("Are you buying something for a child?", ["Children's Products", "Toys"]),
            ("Do you need perfume or self-care products?", ["Perfume"]),
            ("Are you searching for clothing or footwear?", ["Clothing and Footwear"]),
            ("Do you want something for sports or health?", ["Sports and Health"]),
            ("Are you choosing items for hobbies or creativity?", ["Hobbies and Creativity"]),
            ("Do you need products for your car?", ["Auto Products"]),
            ("Do you need something for your pet?", ["Pet Products"]),
            ("Do you want something for the kitchen?", ["Kitchen Products"]),
            ("Do you want something for entertainment—games, consoles, board games?", ["Gaming", "Toys"]),
        ]

        self.process_questionnaire(questions)

    def process_questionnaire(self, questions):
        """FIXED: Reduced boost from 0.5 to 0.15"""
        user = self.current_user

        for question, spheres in questions:
            print("\n" + question)
            print("1. Yes")
            print("2. No")
            answer = self.get_choice(2)

            if answer == 1:
                for sphere in spheres:
                    if sphere in user.sphere_scores:
                        user.sphere_scores[sphere] = min(
                            user.sphere_scores[sphere] + 0.15,
                            MAX_SPHERE_SCORE
                        )
                    else:
                        user.sphere_scores[sphere] = 0.15

        users = self.db.load_users()
        users[user.username] = user.to_dict()
        self.db.save_users(users)

        print("\nYour preferences were updated successfully!")
        input("\nPress Enter to continue...")

    def register(self):
        """User registration"""
        self.clear_screen()
        self.print_header("REGISTRATION")
    
        users = self.db.load_users()
    
        while True:
            username = input("Enter username: ").strip()
            if not username or username.isdigit():
                print("Username cannot be empty or number")
                continue
            if username in users:
                print("Username already exists")
                continue
            break
    
        password = input("Enter password: ").strip()
        if not password:
            print("Password cannot be empty")
            input("\nPress Enter to continue...")
            return
        
        while True:
            try:
                age = int(input("Enter your age: "))
                if age < 1 or age > 120:
                    print("Please enter a valid age")
                    continue
                break
            except ValueError:
                print("Please enter a valid number")
        
        print("\nGender:")
        print("1. Male")
        print("2. Female")
        gender_choice = self.get_choice(2)
        gender = "male" if gender_choice == 1 else "female"
        
        print("\nLocation:")
        print("1. Big City")
        print("2. Small City")
        print("3. Village")
        location_choice = self.get_choice(3)
        if location_choice == 1:
            location = "big_city"
        elif location_choice == 2:
            location = "small_city"
        else:
            location = "village"

        while True:
            try:
                balance = float(input("Enter your balance: "))
                if balance < 0:
                    print("You cannot have a negative balance.")
                    continue
                break
            except ValueError:
                print("Please enter a valid number")

        print(f"Your starting balance is ${balance}")

        user = User(username, password, age, gender, location, balance)

        users[username] = user.to_dict()
        self.db.save_users(users)
        
        print(f"\nRegistration successful! Welcome, {username}!")
        input("\nPress Enter to continue...")
    
    def login(self):
        """User login"""
        self.clear_screen()
        self.print_header("LOGIN")
        
        users = self.db.load_users()
        
        username = input("Enter username: ").strip()
        if username not in users:
            print("Username not found")
            input("\nPress Enter to continue...")
            return
        
        password = input("Enter password: ").strip()
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if users[username]["password_hash"] != password_hash:
            print("Incorrect password")
            input("\nPress Enter to continue...")
            return
        
        self.current_user = User.from_dict(users[username])
        print(f"\nWelcome back, {username}!")
        input("\nPress Enter to continue...")
    
    def logout(self):
        """User logout"""
        users = self.db.load_users()
        users[self.current_user.username] = self.current_user.to_dict()
        self.db.save_users(users)
        
        print(f"\nGoodbye, {self.current_user.username}!")
        self.current_user = None
        input("\nPress Enter to continue...")
    
    def view_recommendations(self):
        """View personalized recommendations"""
        self.clear_screen()
        self.print_header("PERSONALIZED RECOMMENDATIONS")
        
        self.recommendations = RecommendationEngine.get_recommendations(
            self.current_user,
            self.products,
            30
        )
        
        if not self.recommendations:
            print("No recommendations available at the moment.")
            input("\nPress Enter to continue...")
            return
        
        for i, product in enumerate(self.recommendations, 1):
            print(f"{i}. {product['name']}")
            print(f"   Sphere: {product['sphere']} | Type: {product['type']}")
            print(f"   Quality: {product['quality']} | Price: ${product['price']:.2f}")
            print(f"   Delivery: {product['delivery']}")
            print()
        
        print(f"\n{len(self.recommendations) + 1}. Back to Menu")
        
        choice = self.get_choice(len(self.recommendations) + 1)
        
        if choice <= len(self.recommendations):
            self.buy_product(self.recommendations[choice - 1], from_recommendations=True)
    
    def browse_products(self):
        """Browse all products"""
        self.clear_screen()
        self.print_header("ALL PRODUCTS")
        
        spheres = list(set(p["sphere"] for p in self.products))
        spheres.sort()
        
        print("Select a sphere:")
        for i, sphere in enumerate(spheres, 1):
            count = len([p for p in self.products if p["sphere"] == sphere])
            print(f"{i}. {sphere} ({count} products)")
        print(f"\n{len(spheres) + 1}. Back to Menu")
        
        choice = self.get_choice(len(spheres) + 1)
        
        if choice <= len(spheres):
            self.browse_sphere(spheres[choice - 1])
    
    def browse_sphere(self, sphere: str):
        """Browse products in a specific sphere"""
        self.clear_screen()
        self.print_header(f"PRODUCTS - {sphere.upper()}")
        
        sphere_products = [p for p in self.products if p["sphere"] == sphere]
        
        for i, product in enumerate(sphere_products, 1):
            print(f"{i}. {product['name']}")
            print(f"   Type: {product['type']}")
            print(f"   Quality: {product['quality']} | Price: ${product['price']:.2f}")
            print(f"   Delivery: {product['delivery']}")
            print()
        
        print(f"\n{len(sphere_products) + 1}. Back")
        
        choice = self.get_choice(len(sphere_products) + 1)
        
        if choice <= len(sphere_products):
            self.buy_product(sphere_products[choice - 1], from_recommendations=False)
    
    def buy_product(self, product: Dict, from_recommendations: bool):
        """Purchase a product"""
        self.clear_screen()
        self.print_header("PURCHASE PRODUCT")
        
        print(f"Product: {product['name']}")
        print(f"Sphere: {product['sphere']}")
        print(f"Type: {product['type']}")
        print(f"Quality: {product['quality']}")
        print(f"Price: ${product['price']:.2f}")
        print(f"Delivery: {product['delivery']}")
        print(f"\nYour balance: ${self.current_user.balance:.2f}")
        
        print("\n1. Buy")
        print("2. Cancel")
        
        choice = self.get_choice(2)
        
        if choice == 1:
            if product["owner"] == self.current_user.username:
                print("\nYou cannot buy your own product!")
                input("\nPress Enter to continue...")
                return
            
            if self.current_user.balance < product["price"]:
                print("\nInsufficient balance!")
                input("\nPress Enter to continue...")
                return
            
            self.current_user.balance -= product["price"]
            
            if product["owner"] != "system":
                users = self.db.load_users()
                if product["owner"] in users:
                    seller = User.from_dict(users[product["owner"]])
                    seller.balance += product["price"]
                    users[product["owner"]] = seller.to_dict()
                    self.db.save_users(users)
            
            RecommendationEngine.update_profile_after_purchase(
                self.current_user,
                product,
                from_recommendations
            )
            
            transaction = {
                "buyer": self.current_user.username,
                "seller": product["owner"],
                "product": product["name"],
                "price": product["price"],
                "date": datetime.now().isoformat()
            }
            self.db.save_transaction(transaction)
            
            purchase_record = {
                "product_name": product["name"],
                "sphere": product["sphere"],
                "type": product["type"],
                "price": product["price"],
                "quality": product["quality"],
                "seller": product["owner"],
                "date": datetime.now().isoformat()
            }
            self.current_user.purchase_history.append(purchase_record)
            
            users = self.db.load_users()
            users[self.current_user.username] = self.current_user.to_dict()
            self.db.save_users(users)
            
            if "_score" in product:
                del product["_score"]
            if "decay_score" in product:
                del product["decay_score"]
            
            self.products = [p for p in self.products if p["id"] != product["id"]]
            self.db.save_products(self.products)
            
            print("\nPurchase successful!")
            print(f"New balance: ${self.current_user.balance:.2f}")
            input("\nPress Enter to continue...")
    
    def view_profile(self):
        """View user profile"""
        self.clear_screen()
        self.print_header("MY PROFILE")
        
        print(f"Username: {self.current_user.username}")
        print(f"Age: {self.current_user.age}")
        print(f"Gender: {self.current_user.gender}")
        print(f"Location: {self.current_user.location}")
        print(f"Balance: ${self.current_user.balance:.2f}")
        print(f"\nInitial Influence: {self.current_user.initial_influence:.2f}")
        
        print("\nTop 5 Sphere Preferences:")
        sorted_spheres = sorted(
            self.current_user.sphere_scores.items(),
            key=lambda x: x[1],
            reverse=True
        )
        for i, (sphere, score) in enumerate(sorted_spheres[:5], 1):
            print(f"{i}. {sphere}: {score:.2f}")
        
        print("\nCriteria Preferences:")
        print("Quality:")
        for criterion in QUALITY_CRITERIA:
            print(f"  {criterion}: {self.current_user.criteria_scores[criterion]:.2f}")
        print("Price:")
        for criterion in PRICE_CRITERIA:
            print(f"  {criterion}: {self.current_user.criteria_scores[criterion]:.2f}")
        print("Delivery:")
        for criterion in DELIVERY_CRITERIA:
            print(f"  {criterion}: {self.current_user.criteria_scores[criterion]:.2f}")
        
        input("\nPress Enter to continue...")
    
    def view_purchase_history(self):
        """View user's purchase history"""
        self.clear_screen()
        self.print_header("MY PURCHASE HISTORY")
        
        if not self.current_user.purchase_history:
            print("You haven't made any purchases yet.")
            input("\nPress Enter to continue...")
            return
        
        for i, purchase in enumerate(reversed(self.current_user.purchase_history), 1):
            date_obj = datetime.fromisoformat(purchase["date"])
            formatted_date = date_obj.strftime("%Y-%m-%d %H:%M:%S")
            
            print(f"{i}. {purchase['product_name']}")
            print(f"   Sphere: {purchase['sphere']} | Type: {purchase['type']}")
            print(f"   Quality: {purchase['quality']} | Price: ${purchase['price']:.2f}")
            print(f"   Seller: {purchase['seller']}")
            print(f"   Date: {formatted_date}")
            print()
        
        print(f"Total Purchases: {len(self.current_user.purchase_history)}")
        total_spent = sum(p["price"] for p in self.current_user.purchase_history)
        print(f"Total Spent: ${total_spent:.2f}")
        
        input("\nPress Enter to continue...")
    
    def replenish_balance(self):
        """Replenish account balance"""
        self.clear_screen()
        self.print_header("REPLENISH BALANCE")
        
        print(f"Current Balance: ${self.current_user.balance:.2f}\n")
        
        while True:
            try:
                amount = float(input("Enter amount to add: $"))
                if amount <= 0:
                    print("Amount must be positive!")
                    continue
                if amount > 1000000:
                    print("Amount is too large (max $1,000,000)")
                    continue
                break
            except ValueError:
                print("Please enter a valid number")
        
        self.current_user.balance += amount
        
        users = self.db.load_users()
        users[self.current_user.username] = self.current_user.to_dict()
        self.db.save_users(users)
        
        print(f"\n✓ Successfully added ${amount:.2f}")
        print(f"New Balance: ${self.current_user.balance:.2f}")
        input("\nPress Enter to continue...")
    
    def add_product(self):
        """Add a new product to sell"""
        self.clear_screen()
        self.print_header("ADD NEW PRODUCT FOR SALE")
        
        product_name = input("Product name: ").strip()
        if not product_name:
            print("Product name cannot be empty!")
            input("\nPress Enter to continue...")
            return
        
        print("\nChoose a sphere:")
        spheres = sorted(set(p["sphere"] for p in self.products))
        for i, sphere in enumerate(spheres, 1):
            print(f"{i}. {sphere}")
        
        sphere_choice = self.get_choice(len(spheres))
        sphere = spheres[sphere_choice - 1]
        
        print("\nChoose a type:")
        types = SPHERE_TYPES.get(sphere, ["General"])
        
        for i, ptype in enumerate(types, 1):
            print(f"{i}. {ptype}")
        
        type_choice = self.get_choice(len(types))
        product_type = types[type_choice - 1]
        
        print("\nChoose quality:")
        print("1. Premium")
        print("2. Medium")
        print("3. Budget")
        quality_choice = self.get_choice(3)
        quality_map = {1: "premium", 2: "medium", 3: "budget"}
        quality = quality_map[quality_choice]
        
        print("\nChoose price level:")
        print("1. Expensive")
        print("2. Average")
        print("3. Cheap")
        price_choice = self.get_choice(3)
        price_map = {1: "expensive", 2: "average", 3: "cheap"}
        price_level = price_map[price_choice]
        
        print("\nChoose delivery time:")
        print("1. 1 day")
        print("2. 2-3 days")
        print("3. 4+ days")
        delivery_choice = self.get_choice(3)
        delivery_map = {1: "1day", 2: "2-3days", 3: "4+days"}
        delivery = delivery_map[delivery_choice]
        
        while True:
            try:
                price = float(input("\nEnter price ($): "))
                if price <= 0:
                    print("Price must be positive!")
                    continue
                if price > 100000:
                    print("Price is too high (max $100,000)")
                    continue
                break
            except ValueError:
                print("Please enter a valid number")
        
        tags_input = input("\nEnter tags (comma-separated, optional): ").strip()
        tags = [t.strip() for t in tags_input.split(",")] if tags_input else ["user_product"]
        
        new_product_id = max([p["id"] for p in self.products], default=0) + 1
        
        new_product = {
            "id": new_product_id,
            "name": product_name,
            "sphere": sphere,
            "type": product_type,
            "price": price,
            "owner": self.current_user.username,
            "quality": quality,
            "price_level": price_level,
            "delivery": delivery,
            "tags": tags
        }
        
        self.products.append(new_product)
        self.db.save_products(self.products)
        
        print(f"\n✓ Product '{product_name}' successfully listed for sale!")
        print(f"Price: ${price:.2f}")
        input("\nPress Enter to continue...")
    
    def manage_listings(self):
        """Manage your product listings"""
        self.clear_screen()
        self.print_header("MY PRODUCT LISTINGS")
        
        my_products = [p for p in self.products if p["owner"] == self.current_user.username]
        
        if not my_products:
            print("You haven't listed any products for sale.")
            input("\nPress Enter to continue...")
            return
        
        for i, product in enumerate(my_products, 1):
            print(f"{i}. {product['name']}")
            print(f"   Sphere: {product['sphere']} | Type: {product['type']}")
            print(f"   Price: ${product['price']:.2f} | Quality: {product['quality']}")
            print()
        
        print(f"\nTotal listings: {len(my_products)}")
        print(f"\n{len(my_products) + 1}. Back to Menu")
        
        choice = self.get_choice(len(my_products) + 1)
        
        if choice <= len(my_products):
            product_to_withdraw = my_products[choice - 1]
            
            print(f"\nWithdraw '{product_to_withdraw['name']}'?")
            print("1. Yes")
            print("2. No")
            
            confirm = self.get_choice(2)
            if confirm == 1:
                self.products = [p for p in self.products if p["id"] != product_to_withdraw["id"]]
                self.db.save_products(self.products)
                
                print(f"✓ Product withdrawn from sale!")
                input("\nPress Enter to continue...")
            else:
                input("\nPress Enter to continue...")



if __name__ == "__main__":
    app = TerminalInterface()
    app.run()